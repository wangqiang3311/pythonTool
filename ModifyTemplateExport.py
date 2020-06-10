from openpyxl import load_workbook
import os
import shutil
from dynaconf import settings,Validator


class GenerateExportCode:

    """
    自动生成excel导出代码
    """
    def __init__(self):

        self.exportFields=[]


    def modify_cell(self,filePath,min_row=4,min_col=2):
    
        wb = load_workbook(filePath)
        name=wb.sheetnames[0]
        ws=wb[name]

        for row in ws.iter_rows(min_row,min_col):
            for cell in row:
                #if not cell.value:
                    #cell.value="[[Index]]"
                if(cell.value and cell.value.startswith('[[')):
                    field=cell.value.replace('[','').replace('[[','')
                    field=field.replace(']','').replace(']]','')
                    field=field.replace('d.','')
                    exportFields.append(field)
                    cell.value='[[d.{cellValue}]]'.format(cellValue=field)

        wb.save(filePath)

    def re_copyfile(self,srcfile,dstfile):

        """
        复制并修改文件名到另外一个目录
        """

        if not os.path.isfile(srcfile):
            print("%s not exist!" %(srcfile))
        else:
            if os.path.exists(dstfile):
                os.remove(dstfile)
            shutil.copyfile(srcfile,dstfile)
            print("copy %s -> %s"%( srcfile,dstfile))


    def replace_filetag(self,file,old_str,new_str):
        """
        替换文件中的tag标签
        """
        file_data=""
        with open(file,"r",encoding="utf-8") as f:
            for line in f:
                if old_str in line:
                    line=line.replace(old_str,new_str)
                
                file_data+=line
        
        with open(file,"w",encoding='utf-8') as f:
            f.write(file_data)

    def replace_filetags(self,file,tags):
        """
        替换文件中的tag标签
        """
        content=''
        with open(file,"r",encoding="utf-8") as f:
            fileContent=f.read()
            for  key,value in tags.items():
                fileContent= fileContent.replace(key,value)
                
            content=fileContent

        with open(file,"w",encoding='utf-8') as f:
            f.write(content)

   
    def batch_build(self,origin_template_dir):
        """
        批量构建，首先从模板文件夹中获取所有模板，然后遍历
        """

        templateNames=settings.TEMPLATENAMES
        originTemplateDir=settings.ORIGINTEMPLATEDIR
        projectTemplateDir=settings.PROJECTTEMPLATEDIR
        originCodeTemplateDir=settings.ORIGINCODETEMPLATEDIR
        codeDir=settings.CODEDIR
        codeNames=settings.CODENAMES
        reportNames=settings.REPORTNAMES
        templateHistoryNames=settings.TEMPLATEHISTORYNAMES
        reportHistoryNames=settings.REPORTHISTORYNAMES
        DataSourceClassNames=settings.DATASOURCECLASSNAMES
        DataSourceClassHistoryNames=settings.DATASOURCECLASSHISTORYNAMES


        for root,dirs,files in os.walk(origin_template_dir):

            for file in files:

                #修改单元格
                self.modify_cell(os.path.join(root,file))

                #获取在列表中的位置
                index= templateNames.index(file)

                originFile=os.path.join(originTemplateDir,file).replace('\\', '/')
                desFile=os.path.join(projectTemplateDir,templateNames[index]).replace('\\', '/')
             
                #在项目中生成文件以及导出模板
                self.re_copyfile(originFile,desFile)


                codeoriginFile=os.path.join(originCodeTemplateDir,'Template.txt').replace('\\', '/')
                codeDesFile=os.path.join(codeDir,codeNames[index]+".cs").replace('\\', '/')
                self.re_copyfile(codeoriginFile,codeDesFile)
                #替换模板变量
                route="/Iot/"+codeNames[index]

                reportName= reportNames[index].replace('.xlsx','')
                summary=reportName
                reqClass=codeNames[index]
                templateName=templateNames[index]
            
                
                templateNameHistory= templateHistoryNames[index]
                reportNameHistory=reportHistoryNames[index]

                dataSourceClass=DataSourceClassNames[index]
                dataSourceClassHistory=DataSourceClassHistoryNames[index]

                lines=''
                for  f in  self.exportFields:
                    lines+='{f} = w.{f},\n'.format(f=f)

                dataSourceContent=lines
                
                tags={
                    "$w{Route}":route,
                    "$w{Summary}":summary,
                    "$w{RequestClassName}":reqClass,
                    "$w{TemplateName}":templateName,
                    "$w{TemplateNameHistory}":templateNameHistory,
                    "$w{ReportName}":reportName,
                    "$w{ReportNameHistory}":reportNameHistory,
                    "$w{DataSourceClass}":dataSourceClass,
                    "$w{DataSourceClassHistory}":dataSourceClassHistory,
                    "$w_foreach{DataSourceContent}":dataSourceContent,
                }
                
                self.replace_filetags(codeDesFile,tags)

                #清空字段列表
                self.exportFields.clear()
       

        







