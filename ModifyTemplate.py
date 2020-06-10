from openpyxl import load_workbook
import os
import shutil


exportFields=[]

def modifyCell(filePath):
   
    wb = load_workbook(filePath)
    name=wb.sheetnames[0]
    ws=wb[name]

    for row in ws.iter_rows(min_row=4,min_col=2):
        for cell in row:
            #if not cell.value:
                #cell.value="[[Index]]"
            if(cell.value and cell.value.startswith('[[')):
                field=cell.value.replace('[','').replace('[[','')
                field=field.replace(']','').replace(']]','')
                field=field.replace('d.','')
                exportFields.append(field)
                cell.value='[[d.{cellValue}]]'.format(cellValue=field)

    wb.save(filePath)


def re_copyfile(srcfile,dstfile):
    if not os.path.isfile(srcfile):
        print("%s not exist!" %(srcfile))
    else:
        if os.path.exists(dstfile):
            os.remove(dstfile)
        shutil.copyfile(srcfile,dstfile)
        print("copy %s -> %s"%( srcfile,dstfile))


def replaceFileTag(file,old_str,new_str):
    """
    替换文件中的tag标签
    """
    file_data=""
    with open(file,"r",encoding="utf-8") as f:
        for line in f:
            if old_str in line:
                line=line.replace(old_str,new_str)
            
            file_data+=line
    
    with open(file,"w",encoding='utf-8') as f:
        f.write(file_data)

#eg: modifyCell('originTemplate/油井日数据分析.xlsx')

#获取模板文件夹下的所有文件
dirpath='originTemplate'

originTemplateDir='E:\\tools\\Export\\originTemplate'

originCodeTemplateDir='E:\\tools\\Export'

projectTemplateDir='E:\\Project\\jiupaidata\\YCBZ\src\\YCIOT-2020\\Acme\\template'
codeDir='E:\\Project\\jiupaidata\\YCBZ\src\\YCIOT-2020\\YCIOT.Standard\\DataService\\ReportData'


reportNames=['变频最新监测数据.xlsx','电流图最新监测数据.xlsx','功率图最新监测数据.xlsx','功图监测最新数据.xlsx','配水间最新监测数据.xlsx','注水井最新监测数据.xlsx','电参最新监测数据.xlsx']

reportHistoryNames=['变频历史监测数据.xlsx','电流图历史监测数据.xlsx','功率图历史监测数据.xlsx','功图历史监测数据.xlsx','配水间历史监测数据.xlsx','注水井历史监测数据.xlsx','电参历史监测数据.xlsx']

templateNames=['FrequencyCheck.xlsx','CurrentDiagramCheck.xlsx','PowerDiagramCheck.xlsx','GDiagramCheck.xlsx','WaterStationCheck.xlsx','ZWellCheck.xlsx','PowerMetadiagram.xlsx']
templateHistoryNames=['FrequencyHistoryCheck.xlsx','CurrentDiagramHistoryCheck.xlsx','PowerDiagramHistoryCheck.xlsx','GDiagramHistoryCheck.xlsx','WaterStationHistoryCheck.xlsx','ZWellHistoryCheck.xlsx','PowerMetadiagramHistory.xlsx']
codeNames=['FrequencyCheck','CurrentDiagramCheck','PowerDiagramCheck','GDiagramCheck','WaterStationCheck','ZWellCheck','PowerMetadiagram']


DataSourceClassNames=['ViewIotDataOilwellControllerstateLatest','ViewIotDataOilwellCurrentdiagramLatest','ViewIotDataOilWellPowerDiagramLatest','ViewIotDataOilwellIndicatordiagramLatest','ViewIotDataWaterStationLatest','ViewWaterStationWaterWellLatest','ViewIotDataOilwellPowermeterLatest']

DataSourceClassHistoryNames=['ViewIotDataOilwellControllerstate','ViewIotDataOilwellCurrentdiagram','ViewIotDataOilWellPowerDiagram','ViewIotDataOilwellIndicatordiagram','ViewIotDataWaterStation','ViewWaterStationWaterWell','ViewIotDataOilwellPowermeter']


for root,dirs,files in os.walk(dirpath):
    for file in files:

        #修改单元格
        modifyCell(os.path.join(root,file))

        #获取在列表中的位置
        index= templateNames.index(file)
        originFile=os.path.join(originTemplateDir,file).replace('\\', '/')
        desFile=os.path.join(projectTemplateDir,templateNames[index]).replace('\\', '/')
        print('orgin:'+originFile)
        print('des:'+desFile)
        #在项目中生成文件以及导出模板
        re_copyfile(originFile,desFile)


        codeoriginFile=os.path.join(originCodeTemplateDir,'Template.txt').replace('\\', '/')
        codeDesFile=os.path.join(codeDir,codeNames[index]+".cs").replace('\\', '/')
        re_copyfile(codeoriginFile,codeDesFile)
        #替换模板变量
        route="/Iot/"+codeNames[index]

        reportName= reportNames[index].replace('.xlsx','')
        summary=reportName
        reqClass=codeNames[index]
        templateName=templateNames[index]
       
        
        templateNameHistory= templateHistoryNames[index]
        reportNameHistory=reportHistoryNames[index]

        dataSourceClass=DataSourceClassNames[index]
        dataSourceClassHistory=DataSourceClassHistoryNames[index]

        lines=''
        for  f in  exportFields:
            lines+='{f} = w.{f},\n'.format(f=f)

        dataSourceContent=lines
        replaceFileTag(codeDesFile,"$w{Route}",route)
        replaceFileTag(codeDesFile,"$w{Summary}",summary)
        replaceFileTag(codeDesFile,"$w{RequestClassName}",reqClass)
        replaceFileTag(codeDesFile,"$w{TemplateName}",templateName)
        replaceFileTag(codeDesFile,"$w{TemplateNameHistory}",templateNameHistory)
        replaceFileTag(codeDesFile,"$w{ReportName}",reportName)
        replaceFileTag(codeDesFile,"$w{ReportNameHistory}",reportNameHistory)
        replaceFileTag(codeDesFile,"$w{DataSourceClass}",dataSourceClass)
        replaceFileTag(codeDesFile,"$w{DataSourceClassHistory}",dataSourceClassHistory)

        replaceFileTag(codeDesFile," $w_foreach{DataSourceContent}",dataSourceContent)

         #清空字段列表
        exportFields.clear()
       

        







