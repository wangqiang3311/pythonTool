from openpyxl import load_workbook
import os
import shutil


def modifyCell(filePath):
   
    wb = load_workbook(filePath)
    name=wb.sheetnames[0]
    ws=wb[name]

    for row in ws.iter_rows(min_row=4,min_col=2):
        for cell in row:
            if(cell.value):
                field=str(cell.value)
                field=field.replace('[','').replace('[[','')
                field=field.replace(']','').replace(']]','')
                field=field.replace('d.','')
                if field.lower()=='datetime':
                    field='ReturnDateTime'

                if field=='1':
                    field='Index'

                cell.value='[[d.{cellValue}]]'.format(cellValue=field)

    wb.save(filePath)


def re_copyfile(srcfile,dstfile):
    if not os.path.isfile(srcfile):
        print("%s not exist!" %(srcfile))
    else:
        if os.path.exists(dstfile):
            os.remove(dstfile)
        shutil.copyfile(srcfile,dstfile)
        print("copy %s -> %s"%( srcfile,dstfile))


def replaceFileTag(file,old_str,new_str):
    """
    替换文件中的tag标签
    """
    file_data=""
    with open(file,"r",encoding="utf-8") as f:
        for line in f:
            if old_str in line:
                line=line.replace(old_str,new_str)
            
            file_data+=line
    
    with open(file,"w",encoding='utf-8') as f:
        f.write(file_data)

#eg: modifyCell('originTemplate/油井日数据分析.xlsx')

#获取模板文件夹下的所有文件
dirpath='originTemplateHistory'

originTemplateDir='E:\\tools\\Export\\originTemplateHistory'

projectTemplateDir='E:\\Project\\jiupaidata\\YCBZ\src\\YCIOT-2020\\Acme\\template'

templateHistoryNames=['FrequencyHistoryCheck.xlsx','CurrentDiagramHistoryCheck.xlsx','PowerDiagramHistoryCheck.xlsx','GDiagramHistoryCheck.xlsx','WaterStationHistoryCheck.xlsx','ZWellHistoryCheck.xlsx','PowerMetadiagramHistory.xlsx']


for root,dirs,files in os.walk(dirpath):
    for file in files:

        #修改单元格
        modifyCell(os.path.join(root,file))

        #获取在列表中的位置
        index= templateHistoryNames.index(file)
        originFile=os.path.join(originTemplateDir,file).replace('\\', '/')
        desFile=os.path.join(projectTemplateDir,templateHistoryNames[index]).replace('\\', '/')
        print('orgin:'+originFile)
        print('des:'+desFile)
        #在项目中生成文件以及导出模板
        re_copyfile(originFile,desFile)
       

        







